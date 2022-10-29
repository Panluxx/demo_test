# -*- coding: utf-8 -*-
# @Time     : 2021/3/10 20:38
# @Author   : Pan.
# @Email    : 619414118@qq.com
# File      : excel操作.py
# Software  : PyCharm


# 操作 Excel，使用第三方库
import openpyxl

"""
1、 打开excel文件，获取workbook
2、打开表单
3、获取单元格信息
"""

wb = openpyxl.load_workbook('D:\Git\demo_test\cases.xlsx')
print(wb)

sheet = wb['Sheet1']
print(sheet)

data = sheet.cell(1, 1).value
print(data)

a = sheet[1]
print(a)

# 获取所有的数据
# data = sheet.rows
#
# data_list = list(data)[1:]
# print(data_list)
#
# nwe_data = []
# for row in data_list:
#     row_data = []
#     for cell in row:
#         row_data.append(cell.value)
#     nwe_data.append(row_data)
# print(nwe_data)


# 修改
sheet.cell(1, 1).value = 'new_url'

# # 保存
# wb.save('D:\Git\demo_test\cases.xlsx')
#
# # 关闭
# wb.close()
#
# max_row = sheet.max_row
# max_column = sheet.max_column
#
# nwe_data = []
#
# for row in range(2, max_row + 1):
#     row_data = []
#     for column in range(1, max_column + 1):
#         row_data.append(sheet.cell(row, column).value)
#     nwe_data.append(row_data)
# print(nwe_data)
